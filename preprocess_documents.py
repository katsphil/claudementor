#!/usr/bin/env python3
"""
Comprehensive document pre-processing for mentoring reports.
Extracts maximum data while preserving context and narrative.
"""

import json
import re
from pathlib import Path
from typing import Any
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment


def extract_excel_data(file_path: Path) -> dict[str, Any]:
    """Extract comprehensive data from Excel files."""
    workbook = load_workbook(file_path, data_only=False)

    result = {
        "sheet_count": len(workbook.sheetnames),
        "sheets": []
    }

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        ws_data = sheet

        # Extract all data as DataFrame for structured view
        data_rows = []
        for row in ws_data.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data_rows.append(row)

        if data_rows:
            df = pd.DataFrame(data_rows[1:] if len(data_rows) > 1 else [],
                            columns=data_rows[0] if data_rows else None)

            # Convert to dict, handling NaN
            table_data = df.fillna("").to_dict(orient='records')
        else:
            table_data = []

        # Extract comments
        comments = []
        for row in ws_data.iter_rows():
            for cell in row:
                if cell.comment:
                    comments.append({
                        "cell": cell.coordinate,
                        "value": cell.value,
                        "comment": cell.comment.text
                    })

        # Extract formulas
        formulas = []
        for row in ws_data.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.append({
                        "cell": cell.coordinate,
                        "formula": cell.value
                    })

        # Find key financial indicators
        financial_keywords = ['έσοδα', 'έξοδα', 'κέρδος', 'ζημία', 'revenue', 'expenses',
                            'profit', 'loss', 'σύνολο', 'total', 'υπόλοιπο', 'balance']

        key_figures = []
        for row_idx, row in enumerate(ws_data.iter_rows(values_only=True), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell and isinstance(cell, str):
                    cell_lower = cell.lower()
                    if any(keyword in cell_lower for keyword in financial_keywords):
                        # Try to find associated number in next cell or same row
                        try:
                            next_val = row[col_idx] if col_idx < len(row) else None
                            if next_val and isinstance(next_val, (int, float)):
                                key_figures.append({
                                    "label": cell,
                                    "value": next_val,
                                    "location": f"Row {row_idx}"
                                })
                        except:
                            pass

        sheet_info = {
            "name": sheet_name,
            "row_count": ws_data.max_row,
            "column_count": ws_data.max_column,
            "table_data": table_data[:100],  # Limit to first 100 rows
            "has_more_rows": len(data_rows) > 100,
            "comments": comments,
            "formulas": formulas[:50],  # Limit formulas
            "key_figures": key_figures
        }

        result["sheets"].append(sheet_info)

    return result


def extract_pdf_data(file_path: Path) -> dict[str, Any]:
    """Extract comprehensive data from PDF files using Claude Code skills."""
    # Note: This will be called by Claude Code with document-skills plugin
    # For now, return placeholder structure that Claude will fill
    return {
        "extraction_method": "claude_document_skills",
        "requires_processing": True,
        "file_path": str(file_path)
    }


def extract_docx_data(file_path: Path) -> dict[str, Any]:
    """Extract comprehensive data from DOCX files using Claude Code skills."""
    # Note: This will be called by Claude Code with document-skills plugin
    # For now, return placeholder structure that Claude will fill
    return {
        "extraction_method": "claude_document_skills",
        "requires_processing": True,
        "file_path": str(file_path)
    }


def extract_key_excerpts(text: str, keywords: list[str] = None) -> list[str]:
    """Extract key excerpts from text based on keywords and context."""
    if keywords is None:
        keywords = [
            # Greek keywords
            'στόχος', 'στρατηγική', 'πρόβλημα', 'λύση', 'ευκαιρία', 'απειλή',
            'ανάπτυξη', 'επένδυση', 'καινοτομία', 'ανταγωνισμός', 'αγορά',
            # English keywords
            'goal', 'strategy', 'problem', 'solution', 'opportunity', 'threat',
            'growth', 'investment', 'innovation', 'competition', 'market',
            'challenge', 'risk', 'strength', 'weakness'
        ]

    excerpts = []
    sentences = re.split(r'[.!?]\s+', text)

    for sentence in sentences:
        sentence = sentence.strip()
        if len(sentence) < 20:  # Skip very short sentences
            continue

        sentence_lower = sentence.lower()
        if any(keyword in sentence_lower for keyword in keywords):
            # Include surrounding context (up to 300 chars)
            if len(sentence) > 300:
                sentence = sentence[:300] + "..."
            excerpts.append(sentence)

    return excerpts[:20]  # Limit to top 20 excerpts


def assess_document_quality(data: dict) -> list[str]:
    """Assess document quality and professionalism signals."""
    observations = []

    # Check completeness
    if data.get('structured_data'):
        fields = data['structured_data']
        empty_fields = [k for k, v in fields.items() if not v]
        if empty_fields:
            observations.append(f"Missing data in fields: {', '.join(empty_fields[:5])}")

    # Check for key business elements
    text = data.get('full_text', '').lower()

    business_elements = {
        'financial_projections': ['προβλέψεις', 'projections', 'forecast'],
        'swot_analysis': ['swot', 'δυνατά σημεία', 'strengths'],
        'market_analysis': ['αγορά', 'market', 'ανταγωνισμός', 'competition'],
        'action_plan': ['δράσεις', 'action plan', 'ημερομηνία', 'deadline']
    }

    for element, keywords in business_elements.items():
        if any(keyword in text for keyword in keywords):
            observations.append(f"Contains {element.replace('_', ' ')}")
        else:
            observations.append(f"Missing {element.replace('_', ' ')}")

    return observations


def preprocess_document(file_path: Path) -> dict[str, Any]:
    """Main preprocessing function for a single document."""
    file_path = Path(file_path)

    result = {
        "filename": file_path.name,
        "file_type": file_path.suffix.lower(),
        "file_size_kb": file_path.stat().st_size / 1024,
        "modified_date": datetime.fromtimestamp(file_path.stat().st_mtime).isoformat(),
        "extraction_timestamp": datetime.now().isoformat()
    }

    try:
        if file_path.suffix.lower() in ['.xlsx', '.xls']:
            excel_data = extract_excel_data(file_path)
            result["structured_data"] = excel_data
            result["extraction_status"] = "success"

            # Extract text from table data for excerpts
            all_text = []
            for sheet in excel_data.get('sheets', []):
                for row in sheet.get('table_data', []):
                    all_text.extend([str(v) for v in row.values() if v])
            full_text = ' '.join(all_text)
            result["full_text_preview"] = full_text[:5000]  # First 5000 chars
            result["key_excerpts"] = extract_key_excerpts(full_text)

        elif file_path.suffix.lower() == '.pdf':
            result["structured_data"] = extract_pdf_data(file_path)
            result["extraction_status"] = "requires_claude_skills"

        elif file_path.suffix.lower() in ['.docx', '.doc']:
            result["structured_data"] = extract_docx_data(file_path)
            result["extraction_status"] = "requires_claude_skills"

        else:
            result["extraction_status"] = "unsupported_format"

        # Add quality assessment for Excel (others need Claude)
        if result.get("extraction_status") == "success":
            result["quality_observations"] = assess_document_quality(result)

    except Exception as e:
        result["extraction_status"] = "error"
        result["error_message"] = str(e)

    return result


def preprocess_directory(directory: Path, output_path: Path = None) -> dict[str, Any]:
    """Preprocess all business documents in a directory."""
    directory = Path(directory)

    # Discover files
    extensions = ['.pdf', '.xlsx', '.xls', '.docx', '.doc']
    files = []
    for ext in extensions:
        files.extend(directory.glob(f"*{ext}"))
        files.extend(directory.glob(f"**/*{ext}"))

    files = sorted(set(files))

    # Process each file
    results = {
        "preprocessing_timestamp": datetime.now().isoformat(),
        "directory": str(directory),
        "total_files": len(files),
        "files": []
    }

    for file_path in files:
        print(f"Processing: {file_path.name}")
        file_data = preprocess_document(file_path)
        results["files"].append(file_data)

    # Save if output path provided
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print(f"Saved preprocessed data to: {output_path}")

    return results


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: uv run preprocess_documents.py <directory_path> [output_json]")
        sys.exit(1)

    directory = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else directory / "preprocessed_data.json"

    results = preprocess_directory(directory, output_path)
    print(f"\nProcessed {results['total_files']} files")

    # Summary
    statuses = {}
    for file in results['files']:
        status = file['extraction_status']
        statuses[status] = statuses.get(status, 0) + 1

    print("\nExtraction summary:")
    for status, count in statuses.items():
        print(f"  {status}: {count} files")
